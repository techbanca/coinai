#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

    Wrapper around  openpyxl

    Developed by Coin-AI, 2017-2018
    Contact: banca
"""
from openpyxl import load_workbook

def getwb(filename):
    return load_workbook(filename)
    
    
def get_cell(wb,wsname,address):
    ws = wb[wsname]
    return ws[address]

def read_rng_cell(wb,rngname):
    
    name = wb.get_named_range(rngname)
    (sh_name,cell_addr) = [addr for addr in name.destinations][0]
    sh = wb.get_sheet_by_name(sh_name)
    return sh[cell_addr]

def read_rng_value(wb,rngname):
    
    name = wb.get_named_range(rngname)
    (sh_name,cell_addr) = [addr for addr in name.destinations][0]
    sh = wb.get_sheet_by_name(sh_name)
    return sh[cell_addr].value

def read_rng_list(wb,rngname):
    name = wb.get_named_range(rngname)
    addresses = [addr for addr in name.destinations]
    return [adr[0].cell(adr[1]) for adr in addresses]
    
def set_cell_value_sh(sh,row,col, value):
    
    cell = sh.cell(row=row,column=col)
    cell.value = value
    return cell
          
def set_cell_value(wb,rngname,value):
    
    cell = read_rng_cell (wb, rngname)
    cell.value = value
    return cell

