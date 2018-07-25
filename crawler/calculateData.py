
import os, shutil, sys

from openpyxl import load_workbook

current_path = os.path.realpath(__file__)
data_path = current_path.split("crawler")[0] + "data/"

class calculateData():

    def __init__(self):
        self.data_res = []

    def mycopyfile(self, srcfile,dstfile):
        if not os.path.isfile(srcfile):
            print("%s not exist!"%(srcfile))
        else:
            shutil.copyfile(srcfile,dstfile)
            print("copy %s -> %s"%( srcfile,dstfile))

    def call_data(self, data):
        if len(data) > 1:
            pre_item, sec_item = data[:2]
            pre_d, pre_val = pre_item.rstrip("\n").split(",")
            sec_d, sec_val = sec_item.rstrip("\n").split(",")
            val = round(float(float(pre_val) - float(sec_val)) / float(sec_val), 7)
            self.data_res.append([pre_d, val])
            data = data[1:]
            self.call_data(data)

    def read_csv(self, type_name):

        data_path_csv = os.path.join(data_path, "^%s.csv"%type_name)
        
        with open(data_path_csv, 'r') as load_f:
            data_list = load_f.readlines()
            self.call_data(data_list)

    def output_xlsx(self, type_name):

        self.mycopyfile('CoinImportTemplate.xlsx', 'CoinImportTemplate%s.xlsx'%type_name)
        wb = load_workbook('CoinImportTemplate%s.xlsx'%type_name)
        
        sheet_names = wb.sheetnames
        
        w_sheet_0 = wb[sheet_names[0]]
        
        w_sheet_0.cell(row=2,column=2).value = type_name
        w_sheet_0.cell(row=3,column=2).value = type_name

        w_sheet_1 = wb[sheet_names[-1]]
        w_sheet_1.cell(row=2,column=2).value = type_name
        w_sheet_1.cell(row=3,column=2).value = type_name

        self.read_csv(type_name)
        for index, item in enumerate(self.data_res):
            w_sheet_1.cell(row=index + 5,column=1).value = str(item[0]).replace("-","/")
            w_sheet_1.cell(row=index + 5, column=2).value = float(item[1])

        wb.save('CoinImportTemplate%s.xlsx'%type_name)


if __name__ == "__main__":

    if len(sys.argv) < 2:
        print("please input type_name: ")
        exit(-1)

    type_name = str(sys.argv[1])
    # type_name = "tron"
    cal_obj = calculateData()
    cal_obj.output_xlsx(type_name)


